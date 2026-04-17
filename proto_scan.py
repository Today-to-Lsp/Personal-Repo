#!/usr/bin/env python3
# proto_scan.py
# 多协议端口识别 + TLS 细节 + WebSocket(ws/wss) + 默认保存所有结果 + 分类汇总
# - Excel 多工作表（无 openpyxl 自动降级 CSV）
# - --sni auto/none/<domain>；--ws-path 指定 WS 路径
#
# 用法示例：
#   python3 proto_scan.py 114.114.114.114 "12920-12924,80,443" --sni auto --timeout 3 --workers 64
#   python3 proto_scan.py 114.114.114.114 "80-90,443,8080" --out scan_results

import argparse, socket, ssl, sys, re, csv, os, base64, hashlib, tempfile
from concurrent.futures import ThreadPoolExecutor, as_completed
from collections import defaultdict, Counter

# ===================== 通用工具 =====================
def parse_port_spec(spec: str):
    ports = set()
    for part in spec.split(","):
        part = part.strip()
        if not part:
            continue
        m = re.fullmatch(r"(\d+)-(\d+)", part)
        if m:
            a, b = int(m.group(1)), int(m.group(2))
            step = 1 if a <= b else -1
            ports.update(range(a, b + step, step))
        elif part.isdigit():
            ports.add(int(part))
    return sorted(p for p in ports if 1 <= p <= 65535)

def tcp_connect(host: str, port: int, timeout: float):
    s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    s.settimeout(timeout)
    s.connect((host, port))
    return s

def recv_some(sock: socket.socket, timeout: float, max_bytes: int = 65536):
    sock.settimeout(timeout)
    try:
        return sock.recv(max_bytes)
    except Exception:
        return b""

def recv_until(sock: socket.socket, timeout: float, max_bytes: int = 65536):
    sock.settimeout(timeout)
    data = b""
    try:
        while len(data) < max_bytes:
            chunk = sock.recv(8192)
            if not chunk:
                break
            data += chunk
    except Exception:
        pass
    return data

def parse_http_status_line(raw: bytes):
    try:
        head = raw.split(b"\r\n", 1)[0].decode("iso-8859-1", errors="ignore")
        m = re.search(r"HTTP/\d(?:\.\d)?\s+(\d{3})", head)
        return m.group(1) if m else "-"
    except Exception:
        return "-"

def parse_status_and_headers(raw: bytes):
    try:
        header_block = raw.split(b"\r\n\r\n", 1)[0].decode("iso-8859-1", errors="ignore")
        lines = header_block.split("\r\n")
        status = lines[0] if lines else ""
        m = re.search(r"HTTP/\d(?:\.\d)?\s+(\d{3})", status)
        code = m.group(1) if m else "-"
        headers = {}
        for ln in lines[1:]:
            if ":" in ln:
                k, v = ln.split(":", 1)
                headers[k.strip().lower()] = v.strip()
        return code, headers
    except Exception:
        return "-", {}

def decode_cert_from_der(der_bytes: bytes):
    if not der_bytes:
        return {}
    pem = ssl.DER_cert_to_PEM_cert(der_bytes)
    try:
        import _ssl  # type: ignore
        with tempfile.NamedTemporaryFile("w+", delete=True) as f:
            f.write(pem); f.flush()
            return _ssl._test_decode_cert(f.name)  # dict: subject/issuer/notBefore/notAfter/subjectAltName
    except Exception:
        return {}

def extract_cn_issuer_validity(cert_dict: dict):
    def _get_cn():
        try:
            for rdn in cert_dict.get("subject", []):
                for k, v in rdn:
                    if k == "commonName":
                        return v
        except Exception:
            pass
        return "-"
    def _issuer_short():
        parts = []
        try:
            for rdn in cert_dict.get("issuer", []):
                for k, v in rdn:
                    if k in ("organizationName","commonName","countryName"):
                        parts.append(v)
        except Exception:
            return "-"
        return " / ".join(parts) if parts else "-"
    cn = _get_cn()
    issuer = _issuer_short()
    nb = cert_dict.get("notBefore", "-")
    na = cert_dict.get("notAfter", "-")
    return cn, issuer, f"{nb} ~ {na}"

def extract_san_dns(cert_dict: dict):
    out = []
    try:
        for typ, val in cert_dict.get("subjectAltName", []):
            if typ == "DNS":
                out.append(val)
    except Exception:
        pass
    seen = set(); res = []
    for d in out:
        if d not in seen:
            seen.add(d); res.append(d)
    return res

def is_ip_like(s: str):
    return bool(re.fullmatch(r"\d+\.\d+\.\d+\.\d+", s))

# ===================== 协议探测（非 WS） =====================
def detect_ssh(host: str, port: int, timeout: float):
    try:
        s = tcp_connect(host, port, timeout)
        bnr = recv_some(s, timeout, 256); s.close()
        if bnr.startswith(b"SSH-"):
            line = bnr.split(b"\r\n", 1)[0].decode("ascii", errors="ignore")
            return True, line
        return False, "-"
    except Exception:
        return False, "-"

def detect_socks5(host: str, port: int, timeout: float):
    try:
        s = tcp_connect(host, port, timeout)
        s.sendall(b"\x05\x01\x00")
        resp = recv_some(s, timeout, 2); s.close()
        if len(resp) == 2 and resp[0] == 0x05:
            return True, f"0x{resp[1]:02X}"
        return False, "-"
    except Exception:
        return False, "-"

def detect_redis(host: str, port: int, timeout: float):
    try:
        s = tcp_connect(host, port, timeout)
        s.sendall(b"PING\r\n")
        resp = recv_some(s, timeout, 16); s.close()
        return resp.startswith(b"+PONG")
    except Exception:
        return False

def detect_postgres(host: str, port: int, timeout: float):
    try:
        s = tcp_connect(host, port, timeout)
        s.sendall(b"\x00\x00\x00\x08\x04\xd2\x16\x2f")
        resp = recv_some(s, timeout, 1); s.close()
        if resp in (b"S", b"N"):
            return True, resp.decode("ascii")
        return False, "-"
    except Exception:
        return False, "-"

def detect_mysql(host: str, port: int, timeout: float):
    try:
        s = tcp_connect(host, port, timeout)
        data = recv_some(s, timeout, 128); s.close()
        return bool(data and data[0] == 0x0A)
    except Exception:
        return False

def https_tls_probe(host: str, port: int, timeout: float, sni_name, alpn_list: list[str]):
    # sni_name: str 或 None（None 表示禁用 SNI）
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    if alpn_list:
        try: ctx.set_alpn_protocols(alpn_list)
        except NotImplementedError: pass
    try:
        s = tcp_connect(host, port, timeout)
        tls = ctx.wrap_socket(s, server_hostname=(sni_name if sni_name else None)); tls.settimeout(timeout)
        tls_ver = tls.version() or "-"
        cipher = (tls.cipher() or ("-", "", 0))[0]
        try: alpn = tls.selected_alpn_protocol() or "-"
        except Exception: alpn = "-"
        cert_der = b""
        try: cert_der = tls.getpeercert(binary_form=True) or b""
        except Exception: cert_der = b""
        cert_info = decode_cert_from_der(cert_der)
        cn, issuer, validity = extract_cn_issuer_validity(cert_info)
        # HEAD / for HTTP/1.1
        code = "-"
        try:
            host_header = sni_name if (sni_name and not is_ip_like(sni_name)) else host
            req = (f"HEAD / HTTP/1.1\r\nHost: {host_header}\r\n"
                   f"User-Agent: proto-scan/1.4\r\nConnection: close\r\n\r\n").encode("ascii")
            tls.sendall(req)
            raw = recv_until(tls, timeout)
            code = parse_http_status_line(raw)
        except Exception:
            pass
        try: tls.close()
        except Exception: pass
        return True, tls_ver, cipher, alpn, code, cn, issuer, validity, cert_info
    except Exception:
        return False, "-", "-", "-", "-", "-", "-", "-", {}

def http_head_plain(host: str, port: int, timeout: float, host_header: str):
    try:
        s = tcp_connect(host, port, timeout)
        req = (f"HEAD / HTTP/1.1\r\nHost: {host_header}\r\n"
               f"User-Agent: proto-scan/1.4\r\nConnection: close\r\n\r\n").encode("ascii")
        s.sendall(req)
        raw = recv_until(s, timeout); s.close()
        code = parse_http_status_line(raw)
        if re.fullmatch(r"\d{3}", code or ""): return True, code
        return False, "-"
    except Exception:
        return False, "-"

def likely_encrypted_non_tls(host: str, port: int, timeout: float):
    try:
        s = tcp_connect(host, port, timeout)
        first = recv_some(s, timeout, 64); s.close()
        return (not first)
    except Exception:
        return False

# ===================== WebSocket 检测 =====================
GUID = "258EAFA5-E914-47DA-95CA-C5AB0DC85B11"
def _ws_build_key(): return base64.b64encode(os.urandom(16)).decode().strip()
def _ws_accept(key_b64: str):
    raw = base64.b64decode(key_b64.encode()); sha = hashlib.sha1(raw + GUID.encode()).digest()
    return base64.b64encode(sha).decode().strip()

def ws_upgrade_plain(host: str, port: int, timeout: float, host_header: str, path: str):
    try:
        s = tcp_connect(host, port, timeout)
        k = _ws_build_key(); exp = _ws_accept(k)
        req = (f"GET {path} HTTP/1.1\r\nHost: {host_header}\r\nUpgrade: websocket\r\n"
               f"Connection: Upgrade\r\nSec-WebSocket-Key: {k}\r\nSec-WebSocket-Version: 13\r\n"
               f"User-Agent: proto-scan/ws/1.0\r\n\r\n").encode("ascii")
        s.sendall(req)
        raw = recv_until(s, timeout); s.close()
        code, headers = parse_status_and_headers(raw)
        acc = headers.get("sec-websocket-accept", ""); upg = headers.get("upgrade", ""); conn = headers.get("connection", "")
        if code == "101" and acc == exp and "websocket" in upg.lower() and "upgrade" in conn.lower():
            return True, code
        return False, code if code != "-" else "-"
    except Exception:
        return False, "-"

def wss_upgrade_tls(host: str, port: int, timeout: float, sni_name, path: str):
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    try:
        try: ctx.set_alpn_protocols(["http/1.1"])
        except NotImplementedError: pass
        s = tcp_connect(host, port, timeout)
        tls = ctx.wrap_socket(s, server_hostname=(sni_name if sni_name else None)); tls.settimeout(timeout)
        tls_ver = tls.version() or "-"
        cipher = (tls.cipher() or ("-", "", 0))[0]
        try: alpn = tls.selected_alpn_protocol() or "-"
        except Exception: alpn = "-"
        cert_der = b""
        try: cert_der = tls.getpeercert(binary_form=True) or b""
        except Exception: cert_der = b""
        cert_info = decode_cert_from_der(cert_der)
        cn, issuer, validity = extract_cn_issuer_validity(cert_info)
        k = _ws_build_key(); exp = _ws_accept(k)
        host_header = sni_name if (sni_name and not is_ip_like(sni_name)) else host
        req = (f"GET {path} HTTP/1.1\r\nHost: {host_header}\r\nUpgrade: websocket\r\n"
               f"Connection: Upgrade\r\nSec-WebSocket-Key: {k}\r\nSec-WebSocket-Version: 13\r\n"
               f"User-Agent: proto-scan/wss/1.0\r\n\r\n").encode("ascii")
        tls.sendall(req)
        raw = recv_until(tls, timeout)
        try: tls.close()
        except Exception: pass
        code, headers = parse_status_and_headers(raw)
        acc = headers.get("sec-websocket-accept", ""); upg = headers.get("upgrade", ""); conn = headers.get("connection", "")
        ok = (code == "101" and acc == exp and "websocket" in upg.lower() and "upgrade" in conn.lower())
        if ok:
            return True, code, tls_ver, cipher, alpn, cn, issuer, validity, cert_info
        return False, code if code != "-" else "-", tls_ver, cipher, alpn, cn, issuer, validity, cert_info
    except Exception:
        return False, "-", "-", "-", "-", "-", "-", "-", {}

# ===================== SNI 自动模式 =====================
def sni_candidates_from_cert(cert_info: dict, cn_hint: str|None):
    cands = []
    for d in extract_san_dns(cert_info):
        if d and not is_ip_like(d):
            cands.append(d)
    if cn_hint and not is_ip_like(cn_hint):
        cands.append(cn_hint)
    seen = set(); out = []
    for d in cands:
        if d not in seen:
            seen.add(d); out.append(d)
    return out[:10] if len(out) > 10 else out

# ===================== 扫描核心 =====================
def scan_one(host: str, port: int, timeout: float, sni_mode: str, sni_value: str|None, ws_path: str):
    # 返回 dict，字段见 rows_all 规范
    # 1) 优先识别非 Web 协议
    ok, bnr = detect_ssh(host, port, timeout)
    if ok: return row(host, port, "SSH", bnr)

    if detect_redis(host, port, timeout):
        return row(host, port, "REDIS", "PONG")

    okpg, resp = detect_postgres(host, port, timeout)
    if okpg:
        return row(host, port, "PostgreSQL", f"SSL={'yes' if resp=='S' else 'no'}")

    if detect_mysql(host, port, timeout):
        return row(host, port, "MySQL", "Handshake")

    oks5, method = detect_socks5(host, port, timeout)
    if oks5:
        return row(host, port, "SOCKS5", f"method={method}")

    # 2) SNI 模式分支
    if sni_mode == "auto":
        ok0, tls0, cph0, alpn0, code0, cn0, iss0, val0, cert0 = https_tls_probe(host, port, timeout, None, ["http/1.1"])
        auto_cands = sni_candidates_from_cert(cert0, cn0 if cn0 != "-" else None)

        best_tls_other = None
        for name in auto_cands:
            ok_wss, code_wss, tlsv, ciph, alpn, cn, issuer, validity, _ = wss_upgrade_tls(host, port, timeout, name, ws_path)
            if ok_wss:
                return row_tls(host, port, "WSS", code_wss, tlsv, ciph, alpn, cn, issuer, validity, used_host=name)

            ok_tls, tlsv2, ciph2, alpn1, code_https, cn2, issuer2, validity2, _ = https_tls_probe(
                host, port, timeout, name, ["http/1.1"]
            )
            if ok_tls and re.fullmatch(r"\d{3}", code_https or ""):
                return row_tls(host, port, "HTTPS", code_https, tlsv2, ciph2, alpn1, cn2, issuer2, validity2, used_host=name)
            if ok_tls and best_tls_other is None:
                best_tls_other = row_tls(host, port, "TLS-Other", "-", tlsv2, ciph2, alpn1, cn2, issuer2, validity2, used_host=name)

        if best_tls_other is not None:
            return best_tls_other

        ok_ws, code_ws = ws_upgrade_plain(host, port, timeout, host, ws_path)
        if ok_ws:
            return row(host, port, "WS", code_ws)

        ok_http, code_http = http_head_plain(host, port, timeout, host)
        if ok_http:
            return row(host, port, "HTTP", code_http)

        if likely_encrypted_non_tls(host, port, timeout):
            return row(host, port, "Likely-Encrypted(non-TLS)", "-")
        return row(host, port, "-", "-")

    else:
        sni_name = sni_value  # None=禁用SNI；域名=固定SNI
        ok_wss, code_wss, tlsv, ciph, alpn, cn, issuer, validity, _ = wss_upgrade_tls(host, port, timeout, sni_name, ws_path)
        if ok_wss:
            used_host = sni_name if sni_name else host
            return row_tls(host, port, "WSS", code_wss, tlsv, ciph, alpn, cn, issuer, validity, used_host=used_host)

        ok_tls, tlsv2, ciph2, alpn1, code_https, cn2, issuer2, validity2, _ = https_tls_probe(
            host, port, timeout, sni_name, ["http/1.1"]
        )
        if ok_tls and re.fullmatch(r"\d{3}", code_https or ""):
            used_host = sni_name if sni_name else host
            return row_tls(host, port, "HTTPS", code_https, tlsv2, ciph2, alpn1, cn2, issuer2, validity2, used_host=used_host)
        if ok_tls:
            used_host = sni_name if sni_name else host
            return row_tls(host, port, "TLS-Other", "-", tlsv2, ciph2, alpn1, cn2, issuer2, validity2, used_host=used_host)

        host_header = sni_name if (sni_name and not is_ip_like(sni_name)) else host
        ok_ws, code_ws = ws_upgrade_plain(host, port, timeout, host_header, ws_path)
        if ok_ws:
            return row(host, port, "WS", code_ws, used_host=host_header)

        ok_http, code_http = http_head_plain(host, port, timeout, host_header)
        if ok_http:
            return row(host, port, "HTTP", code_http, used_host=host_header)

        if likely_encrypted_non_tls(host, port, timeout):
            return row(host, port, "Likely-Encrypted(non-TLS)", "-")
        return row(host, port, "-", "-")

# 结果行生成（统一字段）
def row(host, port, proto, info, used_host=None):
    return {
        "host": host, "port": port, "proto": proto, "info": info,
        "tls_ver": "-", "cipher": "-", "alpn": "-", "cert_cn": "-", "issuer": "-", "validity": "-",
        "used_host": used_host or host,
        "url": guess_url(proto, used_host or host, port)
    }

def row_tls(host, port, proto, info, tls_ver, cipher, alpn, cn, issuer, validity, used_host=None):
    return {
        "host": host, "port": port, "proto": proto, "info": info,
        "tls_ver": tls_ver, "cipher": cipher, "alpn": alpn, "cert_cn": cn, "issuer": issuer, "validity": validity,
        "used_host": used_host or host,
        "url": guess_url(proto, used_host or host, port)
    }

def guess_url(proto: str, host_header: str, port: int):
    p = proto.upper()
    if p == "HTTPS": return f"https://{host_header}:{port}/"
    if p == "HTTP":  return f"http://{host_header}:{port}/"
    if p == "WSS":   return f"wss://{host_header}:{port}/"
    if p == "WS":    return f"ws://{host_header}:{port}/"
    if p == "TLS-Other": return f"tls://{host_header}:{port}"
    return "-"

# ===================== 导出 & 汇总 =====================
def export_excel_or_csv(all_rows, out_base: str):
    # all_rows: list of dict
    by_category = split_categories(all_rows)
    summary = build_summary(all_rows)

    # 先尝试 Excel
    ok, msg = export_to_excel(all_rows, by_category, summary, out_base + ".xlsx")
    if ok:
        return True, f"{out_base}.xlsx"
    if msg == "openpyxl_not_installed":
        # 多 CSV
        export_to_csv(all_rows, by_category, summary, out_base)
        return False, "csv"
    else:
        # Excel 其它错误，也导出 CSV
        export_to_csv(all_rows, by_category, summary, out_base)
        return False, "csv"

def split_categories(rows):
    cats = {"http_https": [], "ws_wss": [], "non_web": [], "all": rows}
    for r in rows:
        if r["proto"] in ("HTTP", "HTTPS"):
            cats["http_https"].append(r)
        elif r["proto"] in ("WS", "WSS"):
            cats["ws_wss"].append(r)
        else:
            cats["non_web"].append(r)
    return cats

def build_summary(rows):
    total = len(rows)
    by_proto = Counter(r["proto"] for r in rows)
    tls_rows = [r for r in rows if r["proto"] in ("HTTPS","WSS","TLS-Other") and r["tls_ver"] != "-"]
    by_tls = Counter(r["tls_ver"] for r in tls_rows)
    # 各协议端口清单
    ports_by_proto = defaultdict(list)
    for r in rows:
        ports_by_proto[r["proto"]].append(r["port"])
    for k in ports_by_proto:
        ports_by_proto[k] = sorted(set(ports_by_proto[k]))
    return {
        "total": total,
        "by_proto": dict(by_proto),
        "by_tls_ver": dict(by_tls),
        "ports_by_proto": {k: ports_by_proto[k] for k in sorted(ports_by_proto.keys())}
    }

def export_to_excel(all_rows, cats, summary, xlsx_path: str):
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo

        wb = Workbook()
        ws_sum = wb.active; ws_sum.title = "summary"

        # 写 summary：分两块表
        ws_sum.append(["Metric", "Value"])
        ws_sum.append(["Total Ports", summary["total"]])
        ws_sum.append([])
        ws_sum.append(["By Proto", "Count"])
        for k, v in sorted(summary["by_proto"].items()):
            ws_sum.append([k, v])
        ws_sum.append([])
        ws_sum.append(["By TLS Version (HTTPS/WSS/TLS-Other)", "Count"])
        for k, v in sorted(summary["by_tls_ver"].items()):
            ws_sum.append([k, v])
        ws_sum.append([])
        ws_sum.append(["Proto", "Ports"])
        for k, v in summary["ports_by_proto"].items():
            ws_sum.append([k, ", ".join(str(p) for p in v)])

        # 自动列宽
        for col_idx in range(1, 3):
            max_len = 12
            for row in ws_sum.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    max_len = max(max_len, len(str(cell.value)) if cell.value is not None else 0)
            ws_sum.column_dimensions[get_column_letter(col_idx)].width = min(max_len+2, 80)

        # 其他工作表
        def write_sheet(title, rows):
            ws = wb.create_sheet(title=title)
            header = ["host","port","proto","info","tls_ver","cipher","alpn","cert_cn","issuer","validity","used_host","url"]
            ws.append(header)
            for r in rows:
                ws.append([r.get(k, "") for k in header])
            # 列宽 & 表格样式
            for col_idx, name in enumerate(header, 1):
                max_len = max(len(str(name)), *(len(str(rr.get(name,""))) for rr in rows)) if rows else len(name)
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max(10, max_len+2), 60)
            ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}{len(rows)+1}"
            ws.freeze_panes = "A2"
            if rows:
                tbl = Table(displayName=f"T_{title.replace('-','_')}", ref=f"A1:{get_column_letter(len(header))}{len(rows)+1}")
                tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
                ws.add_table(tbl)

        write_sheet("all", cats["all"])
        write_sheet("http_https", cats["http_https"])
        write_sheet("ws_wss", cats["ws_wss"])
        write_sheet("non_web", cats["non_web"])

        wb.save(xlsx_path)
        return True, xlsx_path
    except ModuleNotFoundError:
        return False, "openpyxl_not_installed"
    except Exception as e:
        return False, f"excel_error: {e}"

def export_to_csv(all_rows, cats, summary, out_base: str):
    # summary
    with open(out_base + "-summary.csv", "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Metric","Value"]); w.writerow(["Total Ports", summary["total"]]); w.writerow([])
        w.writerow(["By Proto","Count"])
        for k, v in sorted(summary["by_proto"].items()): w.writerow([k, v])
        w.writerow([]); w.writerow(["By TLS Version (HTTPS/WSS/TLS-Other)","Count"])
        for k, v in sorted(summary["by_tls_ver"].items()): w.writerow([k, v])
        w.writerow([]); w.writerow(["Proto","Ports"])
        for k, v in summary["ports_by_proto"].items(): w.writerow([k, " ".join(str(p) for p in v)])

    def dump(name, rows):
        header = ["host","port","proto","info","tls_ver","cipher","alpn","cert_cn","issuer","validity","used_host","url"]
        with open(out_base + f"-{name}.csv", "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=header); w.writeheader()
            for r in rows: w.writerow({k: r.get(k, "") for k in header})

    dump("all", cats["all"])
    dump("http_https", cats["http_https"])
    dump("ws_wss", cats["ws_wss"])
    dump("non_web", cats["non_web"])

# ===================== 主流程 =====================
def main():
    ap = argparse.ArgumentParser(description="Multi-protocol scanner with TLS & WebSocket; save ALL results; summarized and exported")
    ap.add_argument("host", help="host or ip")
    ap.add_argument("ports", help='port spec, e.g. "80-90,443,8080"')
    ap.add_argument("--timeout", type=float, default=3.0)
    ap.add_argument("--workers", type=int, default=64)
    ap.add_argument("--sni", type=str, default=None, help="SNI/Host header: 'auto' | 'none' | <domain>")
    ap.add_argument("--ws-path", type=str, default="/", help="WebSocket path (default /)")
    ap.add_argument("--out", type=str, default="scan_results", help="output base filename (default scan_results)")
    args = ap.parse_args()

    # SNI 模式
    s = (args.sni or "").strip().lower()
    if s in ("", "none", "-"):
        sni_mode, sni_value = "fixed", None
    elif s == "auto":
        sni_mode, sni_value = "auto", None
    else:
        sni_mode, sni_value = "fixed", args.sni

    ports = parse_port_spec(args.ports)
    if not ports:
        print("No valid ports.", file=sys.stderr); sys.exit(1)

    # 表头
    header = ["HOST","PORT","PROTO","INFO/CODE","TLS_VER","CIPHER","ALPN","CERT_CN","ISSUER","VALIDITY"]
    print("{:<22} {:<6} {:<24} {:<14} {:<8} {:<24} {:<8} {:<32} {:<28} {}".format(*header))
    print("{:<22} {:<6} {:<24} {:<14} {:<8} {:<24} {:<8} {:<32} {:<28} {}".format(
        "-"*22,"-"*6,"-"*24,"-"*14,"-"*8,"-"*24,"-"*8,"-"*32,"-"*28,"-"*17
    ))

    rows_all = []
    with ThreadPoolExecutor(max_workers=args.workers) as ex:
        futs = [ex.submit(scan_one, args.host, p, args.timeout, sni_mode, sni_value, args.ws_path) for p in ports]
        for fut in as_completed(futs):
            r = fut.result()
            rows_all.append(r)
            print("{:<22} {:<6} {:<24} {:<14} {:<8} {:<24} {:<8} {:<32} {:<28} {}".format(
                r["host"], r["port"], r["proto"], (r["info"] or "-")[:14],
                r["tls_ver"], (r["cipher"] or "-")[:24], r["alpn"] or "-",
                (r["cert_cn"] or "-")[:32], (r["issuer"] or "-")[:28], r["validity"] or "-"
            ))

    # 分类 & 汇总 & 导出
    ok_excel, msg = export_excel_or_csv(rows_all, args.out)

    # 终端汇总
    summary = build_summary(rows_all)
    print("\n========== SUMMARY ==========")
    print(f"Total ports scanned: {summary['total']}")
    print("By protocol:")
    for k, v in sorted(summary["by_proto"].items()):
        print(f"  - {k:24s}: {v}")
    if summary["by_tls_ver"]:
        print("TLS versions (HTTPS/WSS/TLS-Other):")
        for k, v in sorted(summary["by_tls_ver"].items()):
            print(f"  - {k:24s}: {v}")
    print("Ports by protocol:")
    for k, ports in sorted(summary["ports_by_proto"].items()):
        plist = ", ".join(str(p) for p in ports)
        print(f"  - {k:24s}: {plist if plist else '-'}")

    if ok_excel:
        print(f"\n[OK] 全量结果 & 分类汇总已导出：{msg}")
    else:
        print(f"\n[OK] 已导出为多 CSV（基础名：{args.out}-*.csv）。如需 Excel： pip3 install openpyxl 后重跑。")

if __name__ == "__main__":
    main()

